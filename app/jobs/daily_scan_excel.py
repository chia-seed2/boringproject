from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


EXPORT_DIR = Path("exports")
LOG_FILENAME = EXPORT_DIR / "daily_scan_log.xlsx"
SUMMARY_SHEET = "Daily Scan Runs"
STORE_SHEET = "Daily Store Breakdown"


SUMMARY_HEADERS = [
    "Run Timestamp (UTC)",
    "Scan Date",
    "Stores Total",
    "Stores Scanned",
    "Stores Failed",
    "Products Saved",
    "Errors Count",
]

STORE_HEADERS = [
    "Run Timestamp (UTC)",
    "Scan Date",
    "Store ID",
    "Store Name",
    "Domain",
    "Status",
    "Products Saved",
    "Error",
]


def _ensure_workbook() -> Workbook:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    if LOG_FILENAME.exists():
        wb = load_workbook(LOG_FILENAME)
    else:
        wb = Workbook()
        default = wb.active
        wb.remove(default)

    if SUMMARY_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(SUMMARY_SHEET)
        ws.append(SUMMARY_HEADERS)

    if STORE_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(STORE_SHEET)
        ws.append(STORE_HEADERS)

    return wb



def append_daily_scan_log(summary: dict, store_rows: Iterable[dict]) -> str:
    wb = _ensure_workbook()

    now = datetime.now(timezone.utc)
    run_timestamp = now.replace(microsecond=0).isoformat()
    scan_date = now.date().isoformat()

    summary_ws = wb[SUMMARY_SHEET]
    summary_ws.append(
        [
            run_timestamp,
            scan_date,
            summary.get("stores_total", 0),
            summary.get("stores_scanned", 0),
            summary.get("stores_failed", 0),
            summary.get("products_saved", 0),
            len(summary.get("errors", [])),
        ]
    )

    store_ws = wb[STORE_SHEET]
    for row in store_rows:
        store_ws.append(
            [
                run_timestamp,
                scan_date,
                row.get("store_id"),
                row.get("store_name"),
                row.get("domain"),
                row.get("status"),
                row.get("products_saved", 0),
                row.get("error"),
            ]
        )

    wb.save(LOG_FILENAME)
    return str(LOG_FILENAME)
