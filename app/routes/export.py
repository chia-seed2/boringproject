from collections import defaultdict
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends
from fastapi.responses import FileResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.db import get_db
from app.models.product import Product
from app.models.variant import Variant
from app.models.price_snapshot import PriceSnapshot
from app.models.store import Store
from app.models.compliance_result import ComplianceResult

router = APIRouter()


def to_float(value):
    if value is None:
        return None
    try:
        return float(value)
    except Exception:
        return None


@router.get("/export/price-history.xlsx")
def export_price_history_excel(db: Session = Depends(get_db)):
    wb = Workbook()
    wb.remove(wb.active)

    # -----------------------------
    # Load all core data
    # -----------------------------
    stores = {s.id: s for s in db.query(Store).all()}
    products = db.query(Product).all()
    variants = db.query(Variant).all()
    snapshots = (
        db.query(PriceSnapshot)
        .order_by(PriceSnapshot.captured_at.asc())
        .all()
    )
    compliance_results = (
        db.query(ComplianceResult)
        .order_by(ComplianceResult.checked_at.desc())
        .all()
    )

    variant_by_product = {}
    for v in variants:
        if v.product_id not in variant_by_product:
            variant_by_product[v.product_id] = v

    latest_compliance_by_variant = {}
    for row in compliance_results:
        if row.variant_id not in latest_compliance_by_variant:
            latest_compliance_by_variant[row.variant_id] = row

    snapshots_by_variant = defaultdict(list)
    for snap in snapshots:
        snapshots_by_variant[snap.variant_id].append(snap)

    # -----------------------------
    # Sheet 1: Products Summary
    # -----------------------------
    ws = wb.create_sheet("Products Summary")
    ws.append([
        "Store",
        "Product Title",
        "Product URL",
        "SKU",
        "First Seen Price",
        "Current Price",
        "Displayed Compare-at Price",
        "Lowest 30-Day Price",
        "Compliance Status",
        "Last Scanned At",
    ])

    for product in products:
        store = stores.get(product.store_id)
        variant = variant_by_product.get(product.id)
        variant_snaps = snapshots_by_variant.get(variant.id, []) if variant else []

        first_seen_price = None
        current_price = None
        current_compare_at_price = None
        last_scanned_at = None

        if variant_snaps:
            first_seen_price = to_float(variant_snaps[0].price)
            current_price = to_float(variant_snaps[-1].price)
            current_compare_at_price = to_float(variant_snaps[-1].compare_at_price)
            last_scanned_at = str(variant_snaps[-1].captured_at)

        compliance = latest_compliance_by_variant.get(variant.id) if variant else None
        lowest_30_day_price = to_float(compliance.lowest_30_day_price) if compliance else None
        compliance_status = compliance.status if compliance else None

        ws.append([
            store.name if store else None,
            product.title,
            product.url,
            variant.sku if variant else None,
            first_seen_price,
            current_price,
            current_compare_at_price,
            lowest_30_day_price,
            compliance_status,
            last_scanned_at,
        ])

    # -----------------------------
    # Sheet 2: Daily Price History
    # One row per snapshot
    # -----------------------------
    ws = wb.create_sheet("Daily Price History")
    ws.append([
        "Scan Date",
        "Store",
        "Product Title",
        "Product URL",
        "SKU",
        "Price",
        "Compare-at Price",
        "Changed Since Previous Scan",
        "Lowest 30-Day Price (Current Calc)",
    ])

    for product in products:
        store = stores.get(product.store_id)
        variant = variant_by_product.get(product.id)
        if not variant:
            continue

        variant_snaps = snapshots_by_variant.get(variant.id, [])
        compliance = latest_compliance_by_variant.get(variant.id)
        lowest_30_day_price = to_float(compliance.lowest_30_day_price) if compliance else None

        previous_price = None

        for snap in variant_snaps:
            current_price = to_float(snap.price)
            changed = None
            if previous_price is not None:
                changed = "YES" if current_price != previous_price else "NO"

            ws.append([
                str(snap.captured_at.date()) if snap.captured_at else None,
                store.name if store else None,
                product.title,
                product.url,
                variant.sku,
                current_price,
                to_float(snap.compare_at_price),
                changed,
                lowest_30_day_price,
            ])

            previous_price = current_price

    # -----------------------------
    # Sheet 3: Price Timeline
    # Wide format: one row per product, dates across columns
    # -----------------------------
    ws = wb.create_sheet("Price Timeline")

    all_dates = sorted(
        {
            str(s.captured_at.date())
            for s in snapshots
            if s.captured_at is not None
        }
    )

    header = ["Store", "Product Title", "Product URL", "SKU"] + all_dates
    ws.append(header)

    for product in products:
        store = stores.get(product.store_id)
        variant = variant_by_product.get(product.id)
        if not variant:
            continue

        variant_snaps = snapshots_by_variant.get(variant.id, [])

        price_by_date = {}
        for snap in variant_snaps:
            if snap.captured_at is None:
                continue
            day = str(snap.captured_at.date())
            price_by_date[day] = to_float(snap.price)

        row = [
            store.name if store else None,
            product.title,
            product.url,
            variant.sku,
        ]

        for day in all_dates:
            row.append(price_by_date.get(day))

        ws.append(row)

    # -----------------------------
    # Save file
    # -----------------------------
    export_dir = Path("exports")
    export_dir.mkdir(exist_ok=True)

    filename = export_dir / f"price_history_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)

    return FileResponse(
        path=filename,
        filename=filename.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )